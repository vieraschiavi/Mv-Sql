"""
exportar.py — MV SQL NLP
==================================================================
Exportación de resultados a Excel (.xlsx), CSV, PDF y HTML.
Todos devuelven bytes listos para descargar desde la interfaz.
==================================================================
"""

import io
from datetime import datetime


def a_csv(df):
    return df.to_csv(index=False).encode("utf-8-sig")  # BOM: Excel abre acentos bien


def a_excel(df, titulo="Consulta MV SQL NLP", sql=""):
    """Excel con formato: encabezado con estilo, autofiltro y hoja de metadatos."""
    buf = io.BytesIO()
    import pandas as pd
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultado", index=False, startrow=1)
        ws = writer.sheets["Resultado"]
        ws["A1"] = titulo

        from openpyxl.styles import Font, PatternFill
        ws["A1"].font = Font(bold=True, size=13, color="1E3A8A")
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row=2, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
        # ancho de columnas
        for i, col in enumerate(df.columns, start=1):
            ancho = max(len(str(col)), *(len(str(v)) for v in df[col].head(200))) + 2
            ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = min(40, ancho)
        if len(df):
            ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(df.columns)).column_letter}{len(df) + 2}"

        meta = pd.DataFrame({
            "Campo": ["Generado por", "Fecha", "Filas", "SQL"],
            "Valor": ["MV SQL NLP", datetime.now().strftime("%Y-%m-%d %H:%M"),
                      len(df), sql[:1000]],
        })
        meta.to_excel(writer, sheet_name="Info", index=False)
    return buf.getvalue()


def a_pdf(df, titulo="Consulta MV SQL NLP", pregunta="", sql="",
          explicacion="", max_filas=200):
    """PDF profesional con tabla, pregunta y explicación (reportlab)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)

    buf = io.BytesIO()
    pagina = landscape(A4) if len(df.columns) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=pagina,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    estilos = getSampleStyleSheet()
    azul = colors.HexColor("#1e3a8a")
    h1 = ParagraphStyle("h1", parent=estilos["Title"], textColor=azul, fontSize=16)
    normal = estilos["BodyText"]

    partes = [Paragraph(titulo, h1)]
    if pregunta:
        partes += [Paragraph(f"<b>Pregunta:</b> {pregunta}", normal), Spacer(1, 4)]
    if explicacion:
        partes += [Paragraph(f"<b>Análisis:</b> {explicacion}", normal), Spacer(1, 4)]
    partes.append(Paragraph(
        f"<font size=8 color='#64748b'>Generado por MV SQL NLP · "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')} · {len(df)} filas"
        f"{' (mostrando ' + str(max_filas) + ')' if len(df) > max_filas else ''}</font>",
        normal))
    partes.append(Spacer(1, 8))

    datos = [list(df.columns)] + df.head(max_filas).astype(str).values.tolist()
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    partes.append(tabla)

    if sql:
        partes += [Spacer(1, 10),
                   Paragraph("<b>SQL ejecutado:</b>", normal),
                   Paragraph(f"<font face='Courier' size=7>{sql.replace(chr(10), '<br/>')}</font>", normal)]

    doc.build(partes)
    return buf.getvalue()


def a_html(df, titulo="Consulta MV SQL NLP"):
    tabla = df.to_html(index=False, border=0)
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{titulo}</title>
<style>
body{{font-family:system-ui,Segoe UI,Arial;margin:2rem;color:#0f172a}}
h1{{color:#1e3a8a}} table{{border-collapse:collapse;width:100%}}
th{{background:#1e3a8a;color:#fff;padding:.5rem;text-align:left}}
td{{padding:.4rem .5rem;border-bottom:1px solid #e2e8f0}}
tr:nth-child(even) td{{background:#f8fafc}}
</style></head><body><h1>{titulo}</h1>{tabla}
<p style="color:#64748b;font-size:.8rem">Generado por MV SQL NLP</p></body></html>"""
    return html.encode("utf-8")
